#!/usr/bin/env python3
"""Aggregate every alt-text manifest under a tree into one master spreadsheet.

Usage:
    python manifest_to_index.py <dir> [<dir> ...] [--out <path.xlsx|.csv>] [--subtitle "<text>"]

One row per image across all folders. Columns:
    Folder | # | File | Category | Alt chars | Alt text | Long description | Notes

Output format follows --out's extension (.xlsx or .csv). Default: <dir>/alt-text-index.xlsx
when openpyxl is installed, else alt-text-index.csv (CSV needs no dependencies).
"""
import os, sys, csv, glob, argparse

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from manifest_common import (parse_manifest, clean_note, norm_cat, norm_key, project_name)

HEADERS = ["Folder", "#", "File", "Category", "Alt chars", "Alt text", "Long description", "Notes"]

def collect(paths):
    manifests = []
    for p in paths:
        if os.path.isdir(p):
            manifests += glob.glob(os.path.join(p, "**", "*alt-text-manifest.md"), recursive=True)
            manifests += glob.glob(os.path.join(p, "*alt-text-manifest.md"))
        elif p.endswith(".md"):
            manifests.append(p)
    manifests = sorted(set(manifests))
    rows = []
    for m in manifests:
        model = parse_manifest(m)
        folder = project_name(model['title'])
        for r in model['rows']:
            alt = r['alt'].strip()
            long_paras = model['longs'].get(norm_key(r['file'])) or []
            rows.append([
                folder, r['idx'], r['file'], norm_cat(r['category']),
                len(alt), alt, "\n\n".join(long_paras), clean_note(r['notes_raw']),
            ])
    return rows, len(manifests)

def write_csv(rows, out):
    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        w.writerows(rows)

def write_xlsx(rows, out):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Alt text index"
    ws.append(HEADERS)
    head_fill = PatternFill("solid", fgColor="1F4E79")
    for c, _ in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(r)
    widths = {"Folder": 30, "#": 5, "File": 34, "Category": 18,
              "Alt chars": 10, "Alt text": 60, "Long description": 70, "Notes": 50}
    for c, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(h, 20)
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    wb.save(out)

def main():
    ap = argparse.ArgumentParser(description="Aggregate alt-text manifests into one spreadsheet.")
    ap.add_argument("paths", nargs="+", help="dir(s) to recurse (or manifest .md files)")
    ap.add_argument("--out", help="output path; .xlsx or .csv picks the format")
    args = ap.parse_args()

    rows, n = collect(args.paths)
    if not rows:
        sys.exit("No manifests found (looked for *alt-text-manifest.md).")

    out = args.out
    if not out:
        base = args.paths[0] if os.path.isdir(args.paths[0]) else os.path.dirname(args.paths[0]) or "."
        try:
            import openpyxl  # noqa: F401
            out = os.path.join(base, "alt-text-index.xlsx")
        except ImportError:
            out = os.path.join(base, "alt-text-index.csv")

    if out.lower().endswith(".csv"):
        write_csv(rows, out)
    else:
        try:
            write_xlsx(rows, out)
        except ImportError:
            alt = os.path.splitext(out)[0] + ".csv"
            write_csv(rows, alt)
            print(f"openpyxl not installed — wrote CSV instead: {alt}")
            print(f"(for .xlsx: .venv/bin/pip install openpyxl)")
            print(f"\n{len(rows)} images from {n} folder(s).")
            return

    print(f"OK  {len(rows)} images from {n} folder(s)  ->  {out}")

if __name__ == "__main__":
    main()
